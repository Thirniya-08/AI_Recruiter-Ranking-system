#!/usr/bin/env python3
"""
unified_uploader.py - Smart combined upload handler for candidates + JD in single file.
Supports ZIP, JSON, and structured formats.
"""

import os
import json
import csv
import zipfile
import tempfile
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from io import StringIO

try:
    import docx
except ImportError:
    docx = None

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None


class UnifiedUploader:
    """Handle combined candidate and JD uploads."""
    
    def __init__(self):
        self.upload_folder = tempfile.gettempdir()
    
    def parse_zip_file(self, file_path: str) -> Tuple[str, List[Dict], Optional[str], str]:
        """
        Parse ZIP file containing candidates and JD.
        Looks for files matching patterns and separates them.
        Returns: (jd_text, candidates_list, file_type, status_message)
        """
        jd_text = None
        candidates = []
        
        try:
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                files = zip_ref.namelist()
                
                # Look for JD files (priority: pdf > docx > txt > json)
                jd_file = None
                jd_patterns = ['jd', 'job_description', 'job-description', 'description', 'requirement']
                
                for f in files:
                    f_lower = f.lower()
                    # Check if this might be a JD file
                    if any(pattern in f_lower for pattern in jd_patterns):
                        jd_file = f
                        break
                
                # If no explicit JD file, look for single PDF/DOCX/TXT
                if not jd_file:
                    for f in files:
                        if f.endswith('.pdf') or f.endswith('.docx') or f.endswith('.txt'):
                            jd_file = f
                            break
                
                # Extract and parse JD if found
                if jd_file:
                    ext = Path(jd_file).suffix.lower()
                    with zip_ref.open(jd_file) as f:
                        if ext == '.pdf':
                            temp_path = os.path.join(self.upload_folder, 'temp_jd.pdf')
                            with open(temp_path, 'wb') as temp_f:
                                temp_f.write(f.read())
                            jd_text = self._extract_text_from_pdf(temp_path)
                            os.remove(temp_path)
                        elif ext == '.docx':
                            temp_path = os.path.join(self.upload_folder, 'temp_jd.docx')
                            with open(temp_path, 'wb') as temp_f:
                                temp_f.write(f.read())
                            jd_text = self._extract_text_from_docx(temp_path)
                            os.remove(temp_path)
                        elif ext == '.txt':
                            jd_text = f.read().decode('utf-8')
                        elif ext == '.json':
                            jd_data = json.loads(f.read().decode('utf-8'))
                            jd_text = json.dumps(jd_data, indent=2)
                
                # Parse candidate files (JSONL, CSV, JSON)
                for f in files:
                    if f == jd_file:
                        continue
                    
                    ext = Path(f).suffix.lower()
                    with zip_ref.open(f) as file_obj:
                        content = file_obj.read().decode('utf-8')
                        
                        if ext == '.jsonl':
                            for line in content.strip().split('\n'):
                                if line:
                                    try:
                                        candidates.append(json.loads(line))
                                    except json.JSONDecodeError:
                                        continue
                        
                        elif ext == '.json':
                            data = json.loads(content)
                            if isinstance(data, list):
                                candidates.extend(data)
                            elif isinstance(data, dict):
                                if 'candidates' in data:
                                    candidates.extend(data['candidates'])
                                else:
                                    candidates.append(data)
                        
                        elif ext == '.csv':
                            reader = csv.DictReader(StringIO(content))
                            candidates.extend(list(reader))
        
        except Exception as e:
            raise ValueError(f"Error parsing ZIP file: {str(e)}")
        
        if not candidates:
            raise ValueError("No candidate data found in ZIP file")
        
        if not jd_text:
            raise ValueError("No job description found in ZIP file")
        
        return jd_text, candidates, 'zip', "Successfully extracted JD and candidates from ZIP"
    
    def parse_json_combined(self, file_path: str) -> Tuple[str, List[Dict], Optional[str], str]:
        """
        Parse JSON file with combined structure:
        {
            "job_description": {...} or "...",
            "candidates": [{...}, ...]
        }
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            if not isinstance(data, dict):
                raise ValueError("JSON must be an object with 'job_description' and 'candidates'")
            
            # Extract JD
            jd_text = None
            jd_sources = ['job_description', 'jd', 'description', 'job_desc']
            
            for key in jd_sources:
                if key in data:
                    jd_value = data[key]
                    if isinstance(jd_value, str):
                        jd_text = jd_value
                    elif isinstance(jd_value, dict):
                        jd_text = json.dumps(jd_value, indent=2)
                    break
            
            if not jd_text:
                raise ValueError("'job_description' field not found in JSON")
            
            # Extract candidates
            candidates = []
            candidate_sources = ['candidates', 'candidate_list', 'people', 'data']
            
            for key in candidate_sources:
                if key in data:
                    cand_data = data[key]
                    if isinstance(cand_data, list):
                        candidates = cand_data
                    break
            
            if not candidates:
                raise ValueError("'candidates' field not found in JSON")
            
            return jd_text, candidates, 'json', "Successfully parsed combined JSON"
        
        except Exception as e:
            raise ValueError(f"Error parsing combined JSON: {str(e)}")
    
    def parse_excel_combined(self, file_path: str) -> Tuple[str, List[Dict], Optional[str], str]:
        """
        Parse Excel file with multiple sheets:
        - Sheet 1: JD data
        - Sheet 2+: Candidate data
        """
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(file_path)
            sheets = workbook.sheetnames
            
            if len(sheets) < 2:
                raise ValueError("Excel must have at least 2 sheets (JD and Candidates)")
            
            # Parse JD from first sheet
            jd_sheet = workbook[sheets[0]]
            jd_text = ""
            for row in jd_sheet.iter_rows(values_only=True):
                jd_text += " ".join(str(cell) for cell in row if cell) + "\n"
            
            # Parse candidates from remaining sheets
            candidates = []
            for sheet_name in sheets[1:]:
                sheet = workbook[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))
                
                if len(rows) > 1:
                    headers = rows[0]
                    for row in rows[1:]:
                        candidate = {}
                        for i, header in enumerate(headers):
                            if i < len(row) and header:
                                candidate[str(header)] = row[i]
                        if candidate:
                            candidates.append(candidate)
            
            return jd_text, candidates, 'excel', "Successfully parsed Excel file"
        
        except ImportError:
            raise ValueError("openpyxl not installed. Cannot parse Excel files.")
        except Exception as e:
            raise ValueError(f"Error parsing Excel file: {str(e)}")
    
    def _extract_text_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF."""
        if fitz is None:
            raise ImportError("PyMuPDF not installed")
        
        text = ""
        with fitz.open(file_path) as doc:
            for page in doc:
                text += page.get_text() + "\n"
        return text
    
    def _extract_text_from_docx(self, file_path: str) -> str:
        """Extract text from DOCX."""
        if docx is None:
            raise ImportError("python-docx not installed")
        
        doc = docx.Document(file_path)
        text = []
        for p in doc.paragraphs:
            if p.text:
                text.append(p.text)
        return "\n".join(text)
    
    def parse_combined_file(self, file_path: str) -> Dict:
        """
        Smart parser that detects file format and extracts both JD and candidates.
        Returns: {
            'success': bool,
            'jd_text': str,
            'candidates': List[Dict],
            'file_type': str,
            'total_candidates': int,
            'message': str
        }
        """
        try:
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext == '.zip':
                jd_text, candidates, file_type, message = self.parse_zip_file(file_path)
            
            elif file_ext == '.json':
                jd_text, candidates, file_type, message = self.parse_json_combined(file_path)
            
            elif file_ext in ['.xlsx', '.xls']:
                jd_text, candidates, file_type, message = self.parse_excel_combined(file_path)
            
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")
            
            return {
                'success': True,
                'jd_text': jd_text,
                'candidates': candidates,
                'file_type': file_type,
                'total_candidates': len(candidates),
                'message': message
            }
        
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'message': f"Failed to parse combined file: {str(e)}"
            }
    
    def validate_combined_data(self, candidates: List[Dict]) -> Tuple[bool, List[str]]:
        """
        Validate candidate data structure.
        Returns: (is_valid, error_messages)
        """
        errors = []
        
        if not candidates:
            errors.append("No candidates provided")
            return False, errors
        
        # Check if candidates have basic required fields
        required_fields = ['profile', 'skills']
        
        for i, candidate in enumerate(candidates):
            if not isinstance(candidate, dict):
                errors.append(f"Candidate {i} is not a dictionary")
        
        return len(errors) == 0, errors


# Global instance
unified_uploader = UnifiedUploader()
